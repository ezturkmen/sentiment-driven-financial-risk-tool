import requests
import pandas as pd
from transformers import pipeline
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment

# --- 1. CONFIGURATION ---
# Replace with your actual NewsAPI key
API_KEY = 'YOUR_API_KEY_HERE'  # <-- INSERT YOUR NEWSAPI KEY HERE

# --- 2. NEWS FETCHING MODULE ---
def fetch_news(query="Ireland economy"):
    """
    Fetches the latest news headlines using NewsAPI.
    """
    print(f"--- Fetching news for: '{query}' ---")
    url = f'https://newsapi.org/v2/everything?q={query}&language=en&sortBy=publishedAt&apiKey={API_KEY}'
    
    try:
        response = requests.get(url)
        data = response.json()
        
        if 'articles' not in data:
            print("Error: Could not retrieve articles. Check your API Key.")
            return []
            
        # Extract the top 5 headlines
        headlines = [article['title'] for article in data['articles'][:5]]
        return headlines
    except Exception as e:
        print(f"An error occurred during news fetching: {e}")
        return []

# --- 3. PROFESSIONAL EXCEL EXPORT MODULE ---
def save_to_excel(df, risk_score):
    """
    Exports the analysis to a stylized Excel report.
    """
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Dublin_Risk_Report_{timestamp}.xlsx"
    
    # Initialize Excel writer with Openpyxl engine
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Risk Analysis')
    
    workbook  = writer.book
    worksheet = writer.sheets['Risk Analysis']

    # STYLING: Header (Navy Blue Background, White Bold Text)
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # STYLING: Conditional Formatting based on Sentiment
    for row in range(2, len(df) + 2):
        sentiment_cell = worksheet.cell(row=row, column=3) # Column 3 is 'Sentiment'
        val = sentiment_cell.value
        
        if val == 'negative':
            sentiment_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid') # Light Red
        elif val == 'positive':
            sentiment_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid') # Light Green
        else:
            sentiment_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid') # Yellow

    # STYLING: Column Widths
    worksheet.column_dimensions['A'].width = 20 # Timestamp
    worksheet.column_dimensions['B'].width = 65 # Headline
    worksheet.column_dimensions['C'].width = 15 # Sentiment
    worksheet.column_dimensions['D'].width = 15 # Confidence

    writer.close()
    print(f"✨ EXCEL REPORT GENERATED: '{filename}'")

# --- 4. SENTIMENT ANALYSIS & ORCHESTRATION ---
print("--- Loading AI Model (FinBERT)... ---")
# Using FinBERT: Pre-trained on financial vocabulary for higher accuracy
nlp_model = pipeline("sentiment-analysis", model="ProsusAI/finbert")

def run_analysis():
    # Step 1: Get News
    headlines = fetch_news()
    if not headlines:
        return

    # Step 2: Analyze Sentiment
    print("--- Analyzing market sentiment... ---")
    results = nlp_model(headlines)
    
    # Step 3: Create DataFrame
    df = pd.DataFrame({
        'Analysis Timestamp': datetime.now().strftime("%Y-%m-%d %H:%M"),
        'Headline': headlines,
        'Sentiment': [res['label'] for res in results],
        'Confidence': [round(res['score'], 3) for res in results]
    })
    
    # Step 4: Calculate Mathematical Risk Index
    # Definition: Percentage of negative news in the sample
    negative_count = (df['Sentiment'] == 'negative').sum()
    total_risk_index = (negative_count / len(headlines)) * 100
    
    # Step 5: Terminal Output
    print("\n" + "="*60)
    print("   FINANCIAL RISK ASSESSMENT REPORT - DUBLIN")
    print("="*60)
    print(df[['Headline', 'Sentiment', 'Confidence']])
    print("-" * 60)
    print(f"TOTAL RISK INDEX: {total_risk_index}%")
    print("="*60)

    # Step 6: Export to Excel
    save_to_excel(df, total_risk_index)

# --- START EXECUTION ---
if __name__ == "__main__":
    run_analysis()